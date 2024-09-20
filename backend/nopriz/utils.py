import itertools
import logging
import os
import re
import time
from io import BytesIO

import pandas as pd
import pytesseract
import requests
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from PIL import Image

from .models import NoprizFiz, NoprizYr


def generate_combinations_of_replacements(
    s,
    targets=["5", "6", "Б", "А", "О", "4", "У"],
    replacements=["6", "5", "6", "4", "0", "3", "9"],
):

    if len(targets) != len(replacements):
        raise ValueError(
            "Количество элементов в 'targets' и 'replacements' должно совпадать"
        )

    all_combinations = set([s])

    for target, replacement in zip(targets, replacements):
        temp_results = set()
        for variant in all_combinations:
            indices = [i for i, char in enumerate(variant) if char == target]
            if indices:
                for r in range(1, len(indices) + 1):
                    for combination in itertools.combinations(indices, r):
                        temp_list = list(variant)
                        for idx in combination:
                            temp_list[idx] = replacement
                        temp_results.add("".join(temp_list))
        all_combinations.update(temp_results)

    final_results = set()
    for variant in all_combinations:
        if "-" in variant:
            dash_index = variant.index("-")
            part_before_dash = variant[: dash_index + 1]
            part_after_dash = variant[dash_index + 1 :]
            part_after_dash = re.sub(r"\D", "", part_after_dash)
            final_results.add(part_before_dash + part_after_dash)
        else:
            final_results.add(variant)

    return sorted(final_results)


def extract_text_from_image(img_url, max_retries=6, retry_delay=5):
    if not img_url:
        return None
    attempt = 0
    while attempt < max_retries:
        try:
            response = requests.get(img_url, timeout=5)
            response.raise_for_status()
            if not response.content:
                logging.error(f"Empty content received for image {img_url}.")
                return None
            with Image.open(BytesIO(response.content)) as img:
                text = pytesseract.image_to_string(
                    img, config="--psm 6 --oem 3", lang="rus"
                )
                return text.strip()

        except requests.exceptions.RequestException as req_err:
            logging.error(f"Request error for image {img_url}: {req_err}")
        except IOError as io_err:
            logging.error(f"IO error processing image {img_url}: {io_err}")
        except Exception as e:
            logging.error(f"General error processing image {img_url}: {e}")

        attempt += 1
        logging.info(f"Retrying ({attempt}/{max_retries}) after {retry_delay} seconds.")
        time.sleep(retry_delay)

    logging.error(
        f"Failed to extract text from image {img_url} after {max_retries} attempts."
    )
    return None


def fiz_get_type_of_work(columns):
    type_of_work = columns[5].text.strip()
    if len(type_of_work) > 12:
        return type_of_work
    else:
        type_of_work = columns[7].text.strip()
        if len(type_of_work) > 12:
            return type_of_work


class ExcelGenerator:
    def __init__(self, data, filename, title, column_mapping):
        self.data = data
        self.filename = filename
        self.title = title
        self.column_mapping = column_mapping

    def generate_excel(self):
        df = pd.DataFrame(self.data)
        df.rename(columns=self.column_mapping, inplace=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.title

        worksheet.merge_cells("A1:H1")
        header = worksheet["A1"]
        header.value = self.title
        header.font = Font(size=16, bold=True, color="6b4fff")
        header.alignment = Alignment(horizontal="center", vertical="center")

        for idx, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=2, column=idx)
            cell.value = column
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            worksheet.column_dimensions[cell.column_letter].width = 25
            if idx == 6:
                worksheet.column_dimensions[cell.column_letter].width = 120

        for row_idx, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx + 3, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left")

        file_path = os.path.join(settings.MEDIA_ROOT, self.filename)
        workbook.save(file_path)
        return file_path


class NoprizFizExcelGenerator(ExcelGenerator):
    def __init__(self):
        data = NoprizFiz.objects.filter(is_parsed=True).values(
            "id_number",
            "full_name",
            "date_of_inclusion_protocol",
            "date_of_modification",
            "date_of_issue_certificate",
            "type_of_work",
            "date_of_exclusion",
            "status_worker",
        )
        filename = "НОПРИЗ Физ лица.xlsx"
        title = "НОПРИЗ Физ-лица"
        column_mapping = {
            "id_number": "Номер ID",
            "full_name": "ФИО",
            "date_of_inclusion_protocol": "Дата включения по протоколу",
            "date_of_modification": "Дата изменения",
            "date_of_issue_certificate": "Дата выдачи сертификата",
            "type_of_work": "Тип работы",
            "date_of_exclusion": "Дата исключения",
            "status_worker": "Статус",
        }
        super().__init__(data, filename, title, column_mapping)


class NoprizYrExcelGenerator(ExcelGenerator):
    def __init__(self):
        data = NoprizYr.objects.all().values(
            "id_number",
            "name_cpo",
            "status",
            "name_of_the_member_cpo",
            "inn",
            "ogrn",
            "date_of_termination",
            "date_of_registration",
            "director",
        )
        filename = "НОПРИЗ Юр лица.xlsx"
        title = "НОПРИЗ Юр-лица"
        column_mapping = {
            "id_number": "Номер ID",
            "name_cpo": "Название СРО",
            "status": "Статус",
            "name_of_the_member_cpo": "Название члена СРО",
            "inn": "ИНН",
            "ogrn": "ОГРН",
            "date_of_termination": "Дата прекращения",
            "date_of_registration": "Дата регистрации",
            "director": "Директор",
        }
        super().__init__(data, filename, title, column_mapping)
