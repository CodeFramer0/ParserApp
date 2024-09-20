from io import BytesIO

import openpyxl
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import render
from nopriz.models import NoprizFiz
from nostroy.models import NostroySmet
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def index(request):
    return render(request, "index.html")


def generate_excel_nopriz_fiz(request):
    data = NoprizFiz.objects.filter(is_parsed = True).values(
        "id_number",
        "full_name",
        "date_of_inclusion_protocol",
        "date_of_modification",
        "date_of_issue_certificate",
        "type_of_work",
        "date_of_exclusion",
        "status_worker",
    )


    df = pd.DataFrame(data)


    column_mapping = {
        "id_number": "Номер ID",
        "full_name": "ФИО",
        "date_of_inclusion_protocol": "Дата включения по протоколу",
        "date_of_modification": "Дата изменения",
        "date_of_issue_certificate": "Дата выдачи сертификата",
        "type_of_work": "Тип работы",
        "date_of_exclusion": "Дата исключения",
        "status_worker": "Статус",
    }
    df.rename(columns=column_mapping, inplace=True)


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="nopriz_fiz_data.xlsx"'


    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Физ лица"


    worksheet.merge_cells("A1:H1")
    header = worksheet["A1"]
    header.value = "НОПРИЗ Физ-лица"
    header.font = Font(size=14, bold=True)
    header.alignment = Alignment(horizontal="center")


    for idx, column in enumerate(df.columns, 1):
        cell = worksheet.cell(row=2, column=idx)
        cell.value = column
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


    for row_idx, row_data in df.iterrows():
        for col_idx, value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx + 3, column=col_idx, value=value)

    workbook.save(response)

    return response
